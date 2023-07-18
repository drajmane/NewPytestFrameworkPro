import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

path="C:\\Users\\admin\\OneDrive\\Desktop\\jiratestcases.xlsx"


workbook=openpyxl.load_workbook(path)
sheet=workbook.active
#
# for r in range(1,3):
#     for c in range(3,6):
#         sheet.cell(row=r,column=c).value="Welcome"
#
#
# workbook.save(path)

rows=sheet.max_row
column=sheet.max_column

print(rows)
print(column)

for r in range(3,rows+1):
    for c in range(1,column+1):
        print(sheet.cell(row=r,column=c).value,end="     ")


    print()













