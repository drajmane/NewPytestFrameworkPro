import openpyxl

def getrowcount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.active
    rows=sheet.max_row
    return rows

def getcolumncount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.active
    column=sheet.max_column
    return column

def readdata(path,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.active
    return  sheet.cell(row=rownum,column=columnno).value

def writedata(path,sheetname,rownum,columnno,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    sheet.cell(row=rownum,column=columnno).value = data
    workbook.save(path)




